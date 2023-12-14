from distutils.command.config import config
import errno
import cv2
from matplotlib import docstring
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from requests import delete
import dataExtractor
import xlsxwriter
import config
import math
import openpyxl
from PIL import Image

def findCountourAtCurrentPosFromNearestCountour(df, x, y):
    try:
    
        x_range = range(x-10,x+10,1)
        y_range = range(y-5,y+5,1)
        df_row_Filtered = df[df.x.isin(x_range) & df.y.isin(y_range) ] #returns values which are in this range
        #print('df fileterd : ',df_row_Filtered)
        if df_row_Filtered.shape[0] == 1:
            index = df_row_Filtered.index.values[0]
            x,y,h,w = df.iloc[index].x, df.iloc[index].y, df.iloc[index].h, df.iloc[index].w
            return x,y,h,w,True
        else:
            x_range = range(x-15,x+15,1)
            y_range = range(y-10,y+10,1)
            df_row_Wide_Filtered = df[df.x.isin(x_range) & df.y.isin(y_range) ]
            if df_row_Wide_Filtered.shape[0] == 1:
                index = df_row_Wide_Filtered.index.values[0]
                x,y,h,w = df.iloc[index].x, df.iloc[index].y, df.iloc[index].h, df.iloc[index].w
                return x,y,h,w, True
            else:
                print('df_row_Wide_Filtered', df_row_Wide_Filtered)
                return 0,0,0,0,False
            
    except ValueError as ve:
        print("Err while findCountourAtCurrentPosFromNearestCountour", ve)


def testTopCountors(table_image_contour, list):
    count = 0
    for l in list:
        if 400< l['pos']['y'] <600:
            print('count:', count)
            print('l=', l )
            roi = table_image_contour[l['pos']['y']:l['pos']['y'] + l['pos']['h'], l['pos']['x']:l['pos']['x']+ l['pos']['w']]
            cv2.imwrite('images/' + str(count) + ".png", roi)
        


        count = count + 1

# def checkAndMergeCellUsingPandas():
#     try:
#         excelfileName = config.excelFileName
#         imageBaseDir = config.imageBaseDir
#         xl = pd.ExcelFile(imageBaseDir + excelfileName)
#         sheet = xl.sheet_names
#         if xl.sheet_names:
#             sheet = xl.sheet_names[0]
#             df = xl.parse(sheet)
#             dimensions = df.shape
#             rowCountMax= dimensions[0]
#             colCountMax= dimensions[1]
#             for index, row in df.iterrows():
#                 print(index, len(row))

        
#     except Exception as Err:
#         print("Error in Cell Merge:", Err)


def checkAndMergeCellIfany(horizontalRows):
    try:
        lenOfhorizontalRows = len(horizontalRows)
        print("HOrizontal rows length : "+(str)(lenOfhorizontalRows))
        print("horizontal rows element length : "+(str)(horizontalRows[0]['colCount']))
        counterForMergechecking = 0
        mergingPointList = []
        for h in horizontalRows:
            # print('h=\n', h)
            if counterForMergechecking < lenOfhorizontalRows -1:
                if horizontalRows[counterForMergechecking ]['colCount'] != horizontalRows[counterForMergechecking + 1 ]['colCount']:
                    if horizontalRows[counterForMergechecking]['colCount'] > horizontalRows[counterForMergechecking + 1]['colCount']:
                        # Find merge start cell
                        row_top_countours =  horizontalRows[counterForMergechecking]['countoutinRow']
                        row_top_bottom =  horizontalRows[counterForMergechecking + 1]['countoutinRow']
                        rb_col_counter = 0
                        
                        for rt in row_top_countours:
                            for rb in row_top_bottom:
                                ratio = (rb['countourInfo']['wTopLeft'] / rt['countourInfo']['wTopLeft'])
                                ceil_ratio = round(ratio)
                                rb_col_counter = rb_col_counter + 1
                                if ceil_ratio >= 2:
                                    print('ceil_ratio::', ceil_ratio)
                                    # Merge starts here
                                    mergingPoint = {
                                        'row': counterForMergechecking + 1,
                                        'colPos': rb_col_counter,
                                        'NoOfColToBeMerged': ceil_ratio
                                    }
                                    mergingPointList.append(mergingPoint)
                                    del mergingPoint
                        print('Merging found below Next:', horizontalRows[counterForMergechecking + 1])
                        print('H=', h)
                    else:
                        print('Merging found above',  h)
            if counterForMergechecking == lenOfhorizontalRows -1:
                if horizontalRows[counterForMergechecking -1 ]['colCount'] != horizontalRows[counterForMergechecking]['colCount']:
                    print('Merging Found at last Row:', horizontalRows[counterForMergechecking])
            counterForMergechecking = counterForMergechecking + 1
        print('mergingPointList=', mergingPointList)
   
    except Exception as Err:
        print("Error while checking merged Cell::", Err)

def checkAndMergeCellsHorizontally(obj):
    #horizontal data updation
    topRow = obj[0]
    #print('top Row -->')
    #print(topRow)
    #print('--------')
    for i in range(1,len(obj)):
        
        if obj[i]['colCount'] != topRow['colCount']:
            #print('current row : '+(str)(i) +', columns : '+(str)(obj[i]['colCount']))
            currentRowCol = 0
            topRowCol = 0
            while currentRowCol < obj[i]['colCount']:
                #current row
                currentRowCellWidth = obj[i]['countoutinRow'][currentRowCol]['countourInfo']['wTopLeft']
                #top row
                y = topRow['countoutinRow'][topRowCol]['countourInfo']['wTopLeft']
                
                ratio = round(currentRowCellWidth/y,1)
                
               # print('x : '+(str)(currentRowCellWidth)+', y : '+(str)(y)+', ratio : '+(str)(ratio))
                
                if ratio == 1:
                    obj[i]['countoutinRow'][currentRowCol]['HorizontalMerged'] = 1
                    currentRowCol += 1 
                    topRowCol += 1
                    
                else:
                    width = topRow['countoutinRow'][topRowCol]['countourInfo']['wTopLeft']
                   # print(width)
                    count = 1
                    while(ratio != 1):
                        width += topRow['countoutinRow'][topRowCol+count]['countourInfo']['wTopLeft']
                     #   print('top row width : '+(str)(width))
                     #   print('current row width : '+(str)(currentRowCellWidth))
                        ratio = round(currentRowCellWidth/width,1) #round off to one decimal
                     #   print('ratio : '+(str)(ratio))
                        count += 1
                    
                    #print('count --->  '+(str)(count))
                    obj[i]['countoutinRow'][currentRowCol]['HorizontalMerged'] = count
                    #print(obj[i])
                    currentRowCol += 1
                    topRowCol += count 
        else:
            currentRowCol = 0
            while currentRowCol < obj[i]['colCount']:
                obj[i]['countoutinRow'][currentRowCol]['HorizontalMerged'] = 1
                currentRowCol += 1
                
    return obj

def checkAndMergeCellsVertically(obj):
   # print(obj)
   # print('total rows : '+(str)(len(obj)))
    for i in range(1,len(obj)):
        #print('columns in current row : '+(str)(obj[i]['colCount']))
        for j in range(0,obj[i]['colCount']):
            currentCellHeight = obj[i]['countoutinRow'][j]['countourInfo']['hTopLeft']
            baseCellHeight = obj[i]['countoutinRow'][0]['countourInfo']['hTopLeft']
            ratio = round(currentCellHeight/baseCellHeight,1)
            count = 1
            if(ratio != 1):
                height = obj[i]['countoutinRow'][0]['countourInfo']['hTopLeft']
                while ratio != 1:
                    height += obj[i+count]['countoutinRow'][0]['countourInfo']['hTopLeft']
                    ratio = round(currentCellHeight/height,1)
                    count += 1
            obj[i]['countoutinRow'][j]['VerticalMerged'] = count

    return obj
                

def mergeCellsDataFix(obj):
    print('show begins : ')
    my_wb_obj = openpyxl.load_workbook(r'C:\Users\ASEN5\Downloads\OneDrive_1_04-04-2022\Script\images\scpa.xlsx')
    my_sheet_obj = my_wb_obj.active
    cols = my_sheet_obj.max_column
    rows = my_sheet_obj.max_row
    
    #print('cols in sheet : '+(str)(cols))
    #print('rows in sheet : '+(str)(rows))
     
    rows = len(obj)
    print('len1 : '+(str)(rows))
    
    for i in range(1,rows):
        cols = obj[i]['colCount']
        for j in range(1,cols):
            #print('row , col = '+(str)(i)+', '+(str)(j))
            #print('v = '+(str)(obj[i]['countoutinRow'][j]['VerticalMerged']))
            #print('h = '+(str)(obj[i]['countoutinRow'][j]['HorizontalMerged']))
            #print('-----')
            v = obj[i]['countoutinRow'][j]['VerticalMerged']
            h = obj[i]['countoutinRow'][j]['HorizontalMerged']
            
            if v > 1 or h > 1:
                cellData = my_sheet_obj.cell(row = i+1,column = j+1).value
                #print('cell data : '+cellData)
                i1,j1,i2,j2 = (i+1),(j+1),(i+v+1),(j+h+1)
                #duplicating according to metaData
                for x in range(i1,i2):
                    for y in range(j1,j2):
                        my_sheet_obj.cell(row = x, column = y).value = cellData
    
    
                        
    my_wb_obj.save('scpaFinal.xlsx')
        
        
    
def extractTextAndWriteDataInExcel(countourList):
    try:
        excelfileName = config.excelFileName
        imageBaseDir = config.imageBaseDir
        workbook = xlsxwriter.Workbook(imageBaseDir + excelfileName)
        worksheet = workbook.add_worksheet()
        row = 0
        col = 0
        counter = 0
        twoHorizontalRowSelctor = 0
        horizontalRows = []
        OneHorizontalRows = []
        ColumnCount = 0
        print('contour size : '+(str)(len(countourList)))
        for countour in countourList:
            # print('countour::', countour)
            text = dataExtractor.extractDataFromImage(countour['imageFileName'], imageBaseDir)
            worksheet.write(row, col, text)
            # print('row:col:text\n', str(row) + ":" + str(col) +":"+ text)

            if counter < ( len(countourList) - 1 ):
                if (countourList[counter + 1]['countourInfo']['y'] - countourList[counter]['countourInfo']['y'] ) > 10:
                    # print('row=', row, countourList[counter + 1]['countourInfo']['y'] ,countourList[counter]['countourInfo']['y'] )
                    
                    # print('Col Count=', col+1, counter)
                
                    rowMetaData = {
                            'row': row,
                            'colCount': col+1,
                            'countoutinRow': countourList[counter -col: counter+1]
                            }
                    horizontalRows.append(rowMetaData)
                    del rowMetaData
                    row = row + 1
                    col = 0    
                    
                else:
                    col = col + 1
                    
            if counter == (len(countourList) - 1):
                print('last row!!')
                rowMetaData = {
                'row': row,
                'colCount': col + 1,
                'countoutinRow': countourList[counter -col: counter+1]
                        }
                horizontalRows.append(rowMetaData)
                del rowMetaData
                
           
            counter = counter + 1
        workbook.close()
        print("----------------")
        print(horizontalRows)
        print("----------------")
        #checkAndMergeCellIfany(horizontalRows)
        horizontalRows = checkAndMergeCellsHorizontally(horizontalRows)
        horizontalRows = checkAndMergeCellsVertically(horizontalRows)
        mergeCellsDataFix(horizontalRows)
        #print(horizontalRows)
        
    except Exception as err:
        print('Error occurred while Extration and preparing Excel::', err)

#sorts the contours according to table and take cell slices
def findContourAndArrange(countourslist, table_image_contour):
    try:
        imageBaseDir = config.imageBaseDir
        #first contour will be largest rectangle
        df = pd.DataFrame(countourslist[1:])
        pos_df = df['pos'].apply(pd.Series)
        # print(pos_df)
       # x_xmin,y_xmin,w_xmin,h_xmin = pos_df.iloc[pos_df['x'].idxmin()]['x'], pos_df.iloc[pos_df['x'].idxmin()]['y'],pos_df.iloc[pos_df['x'].idxmin()]['w'],pos_df.iloc[pos_df['x'].idxmin()]['h']
       # x_ymin,y_ymin,w_ymin,h_ymin = pos_df.iloc[pos_df['y'].idxmin()]['x'], pos_df.iloc[pos_df['y'].idxmin()]['y'],pos_df.iloc[pos_df['y'].idxmin()]['w'],pos_df.iloc[pos_df['y'].idxmin()]['h']

       # x_xmax,y_xmax,w_xmax,h_xmax = pos_df.iloc[pos_df['x'].idxmax()]['x'], pos_df.iloc[pos_df['x'].idxmax()]['y'],pos_df.iloc[pos_df['x'].idxmax()]['w'],pos_df.iloc[pos_df['x'].idxmax()]['h']
       # x_ymax,y_ymax,w_ymax,h_ymax = pos_df.iloc[pos_df['y'].idxmax()]['x'], pos_df.iloc[pos_df['y'].idxmax()]['y'],pos_df.iloc[pos_df['y'].idxmax()]['w'],pos_df.iloc[pos_df['y'].idxmax()]['h']

      #  xTopLeft = x_xmin
      #  yTopLeft = y_ymin
      #  hTopLeft = h_xmin
      #  wTopLeft = w_xmin
        print(pos_df)
        minx = pos_df.iloc[pos_df['x'].idxmin()]['x']
        print('minimum x : ',minx)
        minxList = pos_df.index[pos_df['x'] == minx].tolist()
        print('minimum value index list x :',minxList)
        miny = pos_df.iloc[pos_df['y'].idxmin()]['y']
        print('minimum y : ',miny)
        minyList = pos_df.index[pos_df['y'] == miny].tolist()
        print('minimum value index list y :',minyList)
        
        instIdx = [value for value in minxList if value in minyList]
        instIdx = instIdx[0] 
        print('common index : ',instIdx)
        
        x_ymin,y_ymin,w_ymin,h_ymin = pos_df.iloc[pos_df['y'].idxmin()]['x'], pos_df.iloc[pos_df['y'].idxmin()]['y'],pos_df.iloc[pos_df['y'].idxmin()]['w'],pos_df.iloc[pos_df['y'].idxmin()]['h']

        x_xmax,y_xmax,w_xmax,h_xmax = pos_df.iloc[pos_df['x'].idxmax()]['x'], pos_df.iloc[pos_df['x'].idxmax()]['y'],pos_df.iloc[pos_df['x'].idxmax()]['w'],pos_df.iloc[pos_df['x'].idxmax()]['h']
        x_ymax,y_ymax,w_ymax,h_ymax = pos_df.iloc[pos_df['y'].idxmax()]['x'], pos_df.iloc[pos_df['y'].idxmax()]['y'],pos_df.iloc[pos_df['y'].idxmax()]['w'],pos_df.iloc[pos_df['y'].idxmax()]['h']
        
        x_xmin,y_xmin,w_xmin,h_xmin = pos_df.iloc[instIdx]['x'], pos_df.iloc[instIdx]['y'],pos_df.iloc[instIdx]['w'],pos_df.iloc[instIdx]['h']
        xTopLeft = x_xmin
        yTopLeft = y_xmin
        hTopLeft = h_xmin
        wTopLeft = w_xmin
        print('xtopLeft : ',xTopLeft)
        print('ytopLeft : ',yTopLeft)
        print('htopLeft : ',hTopLeft)
        print('wtopLeft : ',wTopLeft)
        
        print('x_max : '+(str)(x_xmax))
        print('y_max : '+(str)(y_ymax))
        imageCounter = 0
        xboundary = x_xmin
        yBoundary = y_xmin
        row = 0
        col = 0
        imageCountourList = []
        base = False #to keep first cell for each row
        baseRowCell = {}
        #img = cv2.imread(r'C:\Users\ASEN5\Downloads\image extraction\images\image_table.jpg')
       # img = cv2.putText(img,"xmax and y max", (x_xmax,y_ymax), cv2.FONT_HERSHEY_SIMPLEX, 1, (255,0,0))
        while(xboundary < x_xmax  and yBoundary < y_ymax):
            # print('72 imageCounter::', imageCounter, xTopLeft,yTopLeft,hTopLeft,wTopLeft, xboundary,yBoundary, x_xmin,y_xmin, x_xmax,y_ymax )
            roi = table_image_contour[yTopLeft:yTopLeft+hTopLeft, xTopLeft:xTopLeft+wTopLeft]
            imageFileName = str(imageCounter) + ".png"
            cv2.imwrite(imageBaseDir + imageFileName, roi)
            xboundary = xboundary + wTopLeft
            
            countourInfo = {
                'x':xTopLeft,
                'y':yTopLeft,
                'hTopLeft': hTopLeft,
                'wTopLeft': wTopLeft
            }
            #to add data about base cell of a row
            if not base:
                #print('base set')
                baseRowCell = {
                'x':xTopLeft,
                'y':yTopLeft,
                'hTopLeft': hTopLeft,
                'wTopLeft': wTopLeft
                                }
                base = True
                
            #print(baseRowCell['x'])
            #img = cv2.putText(img,(str)( imageCounter), (xTopLeft+25,yTopLeft+25), cv2.FONT_HERSHEY_SIMPLEX, 1, (255,255,255))
                        
            cellMetaData = {
                'imageCounter': imageCounter,
                'imageFileName': imageFileName,
                'countourInfo':countourInfo
            }
            imageCountourList.append(cellMetaData)
            del countourInfo, cellMetaData

            # find next Countor in this position
            
            # print('xboundary::yBoundary', xboundary, yBoundary )
            # xTopLeft,yTopLeft,hTopLeft,wTopLeft = findCountourAtCurrentPos(pos_df, xTopLeft + wTopLeft, yTopLeft, x_xmax, nextRow = False)
            #print(imageCounter)
            #find contours on x-axis
            if xboundary <= x_xmax:
                xTopLeft,yTopLeft,hTopLeft,wTopLeft, status = findCountourAtCurrentPosFromNearestCountour(pos_df, xTopLeft + wTopLeft, yTopLeft)
                if status == False:
                    #1. if next cell is mergedvertically and already included **1.1
                    xTopLeft,yTopLeft,hTopLeft,wTopLeft, status = findCountourAtCurrentPosFromNearestCountour(pos_df, baseRowCell['x'], baseRowCell['y'] + baseRowCell['hTopLeft'])
                    base = False
                    xboundary = x_xmin
                    if status == False:
                        break
                else:
                    # print('86:', xTopLeft,yTopLeft,hTopLeft,wTopLeft)
                
                    col = col + 1
            #find contours on y-axis
            else:
                #2
                xTopLeft,yTopLeft,hTopLeft,wTopLeft, status = findCountourAtCurrentPosFromNearestCountour(pos_df, baseRowCell['x'], baseRowCell['y']+ baseRowCell['hTopLeft'])
                if status == False:
                    break
                else:
                    # print('90:', imageCounter, xTopLeft,yTopLeft,hTopLeft,wTopLeft)
                    xboundary = x_xmin
                    yBoundary = yBoundary + hTopLeft
                    row = row+1
                    col = 0
                    base = False
                 
            imageCounter = imageCounter + 1
            
       # cv2.imwrite('printed.jpg',img)
        return imageCountourList
            

    except Exception as Err:
        print("Error occcurred while fetching corner counters..", Err)



def set_image_dpi(imageFilePath , imageFileName ):
    im = Image.open(imageFilePath + imageFileName)
    length_x, width_y = im.size
    factor = min(1, float(1024.0 / length_x))
    size = int(factor * length_x), int(factor * width_y)
    im_resized = im.resize(size, Image.ANTIALIAS)
    temp_filename = imageFileName.split(".")[0] + "_300-DPI.png"
    im_resized.save(imageFilePath+temp_filename, dpi=(300, 300))
    return temp_filename



def remove_noise_and_smooth(file_name):
    img = cv2.imread(file_name, 0)
    filtered = cv2.adaptiveThreshold(img.astype(np.uint8), 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, 9, 41)
    kernel = np.ones((1, 1), np.uint8)
    opening = cv2.morphologyEx(filtered, cv2.MORPH_OPEN, kernel)
    closing = cv2.morphologyEx(opening, cv2.MORPH_CLOSE, kernel)
    # img = image_smoothening(img)
    or_image = cv2.bitwise_or(img, closing)
    cv2.imwrite(config.imageBaseDir+'noiseCleanedImage.jpg', or_image)
    return or_image



def applythresh(imageAbsPath):
    # imageAbsPath = config.imageBaseDir + config.inputImage
    print('imageAbsPath', imageAbsPath)
    table_image_contour = cv2.imread(imageAbsPath, 0)
    table_image = cv2.imread(imageAbsPath)
    ret, thresh_value = cv2.threshold(
    table_image_contour, 180, 255, cv2.THRESH_BINARY_INV)
    cv2.imwrite(config.imageBaseDir+'threshImage.jpg', thresh_value)
    kernel = np.ones((5,5),np.uint8)
    dilated_value = cv2.dilate(thresh_value,kernel,iterations = 1)
    return dilated_value


def applyImagePreprocessing():
    dpiImage = set_image_dpi(config.imageBaseDir , config.inputImage)
    #noiseCleannedImage = remove_noise_and_smooth(config.imageBaseDir + config.inputImage)
    # cv2.imwrite(config.imageBaseDir + config.inputImage.split(".")[0] + "_OCRReady.png", noiseCleannedImage)
    table_image_contour = cv2.imread(config.imageBaseDir + dpiImage, 0)
    preprocessedImage = applythresh(config.imageBaseDir + dpiImage)
    return preprocessedImage, table_image_contour



def startExtractionProcess(img_tbl):
    try:

        # table_image_contour = table_skeleton # skeleton of the table is passed
        extract_image = img_tbl #table image is passed
        
        dilated_value, table_image_counter = applyImagePreprocessing()
        extract_image = table_image_counter
        
        cnts, hierarchy = cv2.findContours(
        dilated_value, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
        shapes = []

        for cnt in cnts:
            shape = {}
            x, y, w, h = cv2.boundingRect(cnt)
            approx = cv2.contourArea(cnt)
            shape['area'] = approx
            shape_cords = {}
            shape_cords['x'] = x
            shape_cords['y'] = y
            shape_cords['w'] = w
            shape_cords['h'] = h
            shape['pos'] = shape_cords
            shapes.append(shape)

        
        topSizedContours = sorted(shapes, key=lambda d: d['area'], reverse=True)
        # Consider the best possible case of 6X20 docs for best persoformance
        mostPossibleROIlist = topSizedContours[0:250]
        # testTopCountors(table_image_contour, mostPossibleROIlist)

        # Find 4 cornoers contours
        arrangedContourList = findContourAndArrange(mostPossibleROIlist, extract_image)
        extractTextAndWriteDataInExcel(arrangedContourList)


    except Exception as err:
        print("Error Occurred while Extraction..", err)

if __name__ == "__main__":
    
    try:

        # checkAndMergeCellUsingPandas()
        startExtractionProcess(config.imageBaseDir+config.inputImage)
        print('Saving file')
        
    except Exception as err:
        print("Err=", err)

    